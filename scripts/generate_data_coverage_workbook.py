"""Regenerates Data_Coverage_Dashboard.xlsx (project root) from LIVE data -- janmitra.db and
data/rag_knowledge_base/knowledge_records/ -- every time it's run. Nothing here is cached or
remembered from a prior run; re-run this any time the underlying data changes to get a current
snapshot. This is the "real-time" piece the tracker docs (DATA_COVERAGE_TRACKER.md,
RAG_REAL_VS_SYNTHETIC_RESEARCH_PREP.md) point at for a spreadsheet view instead of a markdown one.

Usage:
    python scripts/generate_data_coverage_workbook.py
"""

from __future__ import annotations

import json
import os
import sqlite3
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "janmitra.db"
RAG_DIR = ROOT / "data" / "rag_knowledge_base" / "knowledge_records"
OUT_PATH = ROOT / "Data_Coverage_Dashboard.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
TITLE_FONT = Font(bold=True, size=14)


def style_header_row(ws: Worksheet, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(ws: Worksheet, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def load_rag_records() -> list[dict]:
    records = []
    for tier in ("verified", "synthetic"):
        base = RAG_DIR / tier
        for path in base.rglob("*.json"):
            data = json.loads(path.read_text(encoding="utf-8"))
            if not isinstance(data, list):
                data = [data]
            for rec in data:
                rec["_tier"] = tier
                records.append(rec)
    return records


def main() -> None:
    con = sqlite3.connect(str(DB_PATH))
    cur = con.cursor()

    cur.execute("SELECT name, code, is_union_territory FROM states ORDER BY name")
    all_states = cur.fetchall()

    cur.execute(
        """
        SELECT s.name, d.name, u.name, w.name, l.name
        FROM districts d
        JOIN states s ON d.state_id = s.id
        LEFT JOIN ulbs u ON u.district_id = d.id
        LEFT JOIN wards w ON w.ulb_id = u.id
        LEFT JOIN localities l ON l.ward_id = w.id
        """
    )
    seeded = {row[0]: row[1:] for row in cur.fetchall()}

    records = load_rag_records()
    cats = ["WASTE_SANITATION", "WATER_DRAINAGE", "ROADS_POTHOLES", "STREETLIGHTS"]
    matrix: dict[str, dict[str, dict[str, int]]] = defaultdict(
        lambda: defaultdict(lambda: {"verified": 0, "synthetic": 0})
    )
    state_totals = defaultdict(lambda: {"verified": 0, "synthetic": 0})
    city_totals: dict[tuple[str, str], dict[str, int]] = defaultdict(
        lambda: {"verified": 0, "synthetic": 0}
    )
    for rec in records:
        st = rec.get("state") or "STATE-WIDE/UNSPECIFIED"
        cat = rec.get("service_category", "UNKNOWN")
        tier = rec["_tier"]
        matrix[st][cat][tier] += 1
        state_totals[st][tier] += 1
        city = rec.get("city") or "(state-wide record)"
        city_totals[(st, city)][tier] += 1

    # ---- worker/complaint ward free-text quality check ----
    clean_wards = {v[2] for v in seeded.values() if v[2]}  # ward name e.g. "Ward 22"
    clean_full_strings = set()
    for state, (dist, ulb, ward, loc) in seeded.items():
        if ward and loc and dist:
            clean_full_strings.add(f"{ward} — {loc}, {dist}")

    cur.execute("SELECT id, full_name, ward FROM users WHERE role='worker' AND ward IS NOT NULL")
    worker_wards = cur.fetchall()

    cur.execute("SELECT ward, COUNT(*) FROM complaints WHERE ward IS NOT NULL GROUP BY ward")
    complaint_ward_counts = cur.fetchall()

    con.close()

    wb = Workbook()

    # ================= Sheet 1: Overview =================
    ws = wb.active
    ws.title = "Overview"
    ws["A1"] = "JanSarthi AI — Data Coverage Dashboard"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Regenerated fresh from janmitra.db + data/rag_knowledge_base/ every time this workbook is built."
    ws["A2"].font = Font(italic=True, size=9, color="666666")

    total_v = sum(v["verified"] for v in state_totals.values())
    total_s = sum(v["synthetic"] for v in state_totals.values())
    states_with_any_rag = sum(1 for v in state_totals.values() if v["verified"] + v["synthetic"] > 0)
    clean_worker_count = sum(1 for _, _, w in worker_wards if w in clean_full_strings)

    overview_rows = [
        ("Metric", "Value"),
        ("States/UTs total (real, official list)", len(all_states)),
        ("States/UTs with any knowledge-base content", states_with_any_rag),
        ("States/UTs with ZERO knowledge-base content", len(all_states) - states_with_any_rag),
        ("Knowledge-base records — Verified (real source)", total_v),
        ("Knowledge-base records — Synthetic (placeholder)", total_s),
        ("Cities with structured District→Locality hierarchy", len(seeded)),
        ("Workers with a ward value set", len(worker_wards)),
        ("Workers whose ward is clean production data", clean_worker_count),
        ("Workers whose ward is leftover test-fixture junk", len(worker_wards) - clean_worker_count),
    ]
    for r, (label, val) in enumerate(overview_rows, start=4):
        ws.cell(row=r, column=1, value=label)
        cell = ws.cell(row=r, column=2, value=val)
        if r == 4:
            continue
        if "ZERO" in label or "junk" in label:
            cell.fill = RED_FILL
        elif "Verified" in label or "clean" in label:
            cell.fill = GREEN_FILL
    ws.cell(row=4, column=1).font = Font(bold=True)
    ws.cell(row=4, column=2).font = Font(bold=True)
    style_header_row(ws, 4, 2)
    autosize(ws, [52, 12])
    ws.freeze_panes = "A5"

    # ================= Sheet 2: State Coverage Matrix =================
    ws2 = wb.create_sheet("State_Coverage_Matrix")
    headers = [
        "State/UT", "District", "Sub-District", "ULB", "Zone", "Ward", "Locality",
        "Waste_Verified", "Waste_Synthetic", "Water_Verified", "Water_Synthetic",
        "Roads_Verified", "Roads_Synthetic", "Streetlights_Verified", "Streetlights_Synthetic",
        "TYPE_A_File_Complaint", "TYPE_B_Answer_Question", "TYPE_C_Check_Status",
    ]
    ws2.append(headers)
    style_header_row(ws2, 1, len(headers))
    for name, code, is_ut in all_states:
        has_loc = name in seeded
        dist, ulb, ward, loc = seeded.get(name, (None, None, None, None))
        row = [
            name, "Yes" if has_loc else "No", "No",
            "Yes" if (has_loc and ulb) else "No", "No",
            "Yes" if (has_loc and ward) else "No", "Yes" if (has_loc and loc) else "No",
        ]
        rag_any = False
        for c in cats:
            v = matrix[name][c]["verified"]
            s = matrix[name][c]["synthetic"]
            if v + s > 0:
                rag_any = True
            row += [v, s]
        row += ["Yes", "Yes" if rag_any else "No", "Yes"]
        ws2.append(row)
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=2, max_col=7):
        for cell in row:
            cell.fill = GREEN_FILL if cell.value == "Yes" else RED_FILL
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=16, max_col=18):
        for cell in row:
            cell.fill = GREEN_FILL if cell.value == "Yes" else RED_FILL
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws2.max_row}"
    ws2.freeze_panes = "A2"
    autosize(ws2, [32, 9, 12, 9, 8, 8, 10] + [10] * 8 + [12, 12, 12])

    # ================= Sheet 3: Location Hierarchy Source Types =================
    ws3 = wb.create_sheet("Location_Hierarchy_Sources")
    ws3.append(["Level", "Name", "Parent", "Source_Type", "What_Source_Type_Means"])
    style_header_row(ws3, 1, 5)
    meaning = {
        "OFFICIAL_NATIONAL_REFERENCE": "Checked live against India.gov.in — fully real, official.",
        "WELL_ESTABLISHED_PUBLIC_GEOGRAPHY": "Common public knowledge (e.g. which state a city is in) — real, not independently re-verified via a live URL this session.",
        "UNVERIFIED_APP_SEED_DATA": "Authored for dev/demo purposes — the ward NUMBER and locality NAME are each real, but this specific ward→locality pairing has NOT been checked against an official ward-delimitation source.",
    }
    con = sqlite3.connect(str(DB_PATH))
    cur = con.cursor()
    cur.execute("SELECT name, source_type FROM states ORDER BY name")
    for name, st in cur.fetchall():
        ws3.append(["State", name, "", st, meaning.get(st, "")])
    cur.execute("SELECT d.name, s.name, d.source_type FROM districts d JOIN states s ON d.state_id=s.id ORDER BY d.name")
    for name, parent, st in cur.fetchall():
        ws3.append(["District", name, parent, st, meaning.get(st, "")])
    cur.execute("SELECT u.name, d.name, u.source_type FROM ulbs u JOIN districts d ON u.district_id=d.id ORDER BY u.name")
    for name, parent, st in cur.fetchall():
        ws3.append(["ULB", name, parent, st, meaning.get(st, "")])
    cur.execute("SELECT w.name, u.name, w.source_type FROM wards w JOIN ulbs u ON w.ulb_id=u.id ORDER BY w.name")
    for name, parent, st in cur.fetchall():
        ws3.append(["Ward", name, parent, st, meaning.get(st, "")])
    cur.execute("SELECT l.name, w.name, l.source_type FROM localities l JOIN wards w ON l.ward_id=w.id ORDER BY l.name")
    for name, parent, st in cur.fetchall():
        ws3.append(["Locality", name, parent, st, meaning.get(st, "")])
    con.close()
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, min_col=4, max_col=4):
        for cell in row:
            if cell.value == "OFFICIAL_NATIONAL_REFERENCE":
                cell.fill = GREEN_FILL
            elif cell.value == "WELL_ESTABLISHED_PUBLIC_GEOGRAPHY":
                cell.fill = GREEN_FILL
            elif cell.value == "UNVERIFIED_APP_SEED_DATA":
                cell.fill = YELLOW_FILL
    ws3.auto_filter.ref = f"A1:E{ws3.max_row}"
    ws3.freeze_panes = "A2"
    autosize(ws3, [10, 42, 22, 32, 90])

    # ================= Sheet 4: Ward Data Quality (mock check) =================
    ws4 = wb.create_sheet("Ward_Data_Quality")
    ws4.append(["Worker_Name", "Ward_Value_Stored", "Status"])
    style_header_row(ws4, 1, 3)
    for _id, fname, ward in sorted(worker_wards, key=lambda r: r[2]):
        status = "Clean (matches seeded hierarchy)" if ward in clean_full_strings else "TEST-FIXTURE JUNK (not a real place)"
        ws4.append([fname, ward, status])
    for row in ws4.iter_rows(min_row=2, max_row=ws4.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.fill = GREEN_FILL if cell.value.startswith("Clean") else RED_FILL
    ws4.auto_filter.ref = f"A1:C{ws4.max_row}"
    ws4.freeze_panes = "A2"
    autosize(ws4, [26, 45, 38])

    ws4b = wb.create_sheet("Complaint_Ward_Text_Quality")
    ws4b.append(["Ward_Value_Stored", "Complaint_Count", "Status"])
    style_header_row(ws4b, 1, 3)
    for ward, count in sorted(complaint_ward_counts, key=lambda r: r[0]):
        status = "Clean (matches seeded hierarchy)" if ward in clean_full_strings else "TEST-FIXTURE JUNK / inconsistent (not a real, canonical place string)"
        ws4b.append([ward, count, status])
    for row in ws4b.iter_rows(min_row=2, max_row=ws4b.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.fill = GREEN_FILL if cell.value.startswith("Clean") else RED_FILL
    ws4b.auto_filter.ref = f"A1:C{ws4b.max_row}"
    ws4b.freeze_panes = "A2"
    autosize(ws4b, [45, 16, 55])

    # ================= Sheet 5: RAG City Breakdown =================
    ws5 = wb.create_sheet("RAG_City_Breakdown")
    ws5.append(["State", "City", "Verified_Records", "Synthetic_Records", "Status"])
    style_header_row(ws5, 1, 5)
    for (st, city), counts in sorted(city_totals.items()):
        v, s = counts["verified"], counts["synthetic"]
        if v == 0 and s > 0:
            status = "100% SYNTHETIC -- 0 real records"
        elif s == 0 and v > 0:
            status = "100% real"
        elif v >= s:
            status = "Mostly real"
        else:
            status = "Mostly synthetic"
        ws5.append([st, city, v, s, status])
    for row in ws5.iter_rows(min_row=2, max_row=ws5.max_row, min_col=5, max_col=5):
        for cell in row:
            if cell.value.startswith("100% SYNTHETIC"):
                cell.fill = RED_FILL
            elif cell.value == "100% real":
                cell.fill = GREEN_FILL
            elif cell.value == "Mostly synthetic":
                cell.fill = YELLOW_FILL
    ws5.auto_filter.ref = f"A1:E{ws5.max_row}"
    ws5.freeze_panes = "A2"
    autosize(ws5, [22, 30, 16, 18, 30])

    wb.save(str(OUT_PATH))
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
